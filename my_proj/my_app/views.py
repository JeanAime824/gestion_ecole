from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Avg, Count, Q
from django.core.paginator import Paginator
from django.db.models.functions import TruncMonth
from .forms import (
    ConnexionForm, ResponsableCreationForm, ClasseForm,
    EleveForm, PaiementForm, NoteForm
)
from .models import Classe, Eleve, Utilisateur, Paiement, Note
import json
import csv
import io
from openpyxl import Workbook
from django.http import HttpResponse

def is_admin(user):
    return user.role == 'administrateur' or user.is_superuser

def connexion_view(request):
    if request.method == "POST":
        form = ConnexionForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                if user.role == "administrateur" or user.is_superuser:
                    return redirect('admin_dashboard')
                elif user.role == "responsable":
                    return redirect('responsable_dashboard')
                else:
                    form.add_error(None, "Le rôle de cet utilisateur n'est pas défini.")
            else:
                form.add_error(None, "Nom d'utilisateur ou mot de passe incorrect.")
    else:
        form = ConnexionForm()
    return render(request, 'login.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def ajouter_responsable(request):
    if request.method == "POST":
        form = ResponsableCreationForm(request.POST)
        if form.is_valid():
            responsable = form.save(commit=False)
            responsable.role = 'responsable'
            responsable.save()
            return redirect('admin_dashboard')
    else:
        form = ResponsableCreationForm()
    return render(request, 'ajouter_responsable.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def liste_responsables(request):
    responsables = Utilisateur.objects.filter(role='responsable')
    return render(request, 'liste_responsables.html', {'responsables': responsables})

@login_required
def admin_dashboard(request):
    if not is_admin(request.user):
        return redirect('responsable_dashboard')

    total_eleves = Eleve.objects.count()
    total_paiements = Paiement.objects.aggregate(Sum('montant'))['montant__sum'] or 0
    total_classes = Classe.objects.count()
    moyenne_generale = Note.objects.aggregate(Avg('valeur'))['valeur__avg'] or 0

    stats_classes = Classe.objects.annotate(nb_eleves=Count('eleves'))
    labels_classes = [c.nom for c in stats_classes]
    data_classes = [c.nb_eleves for c in stats_classes]

    paiements_mensuels = Paiement.objects.annotate(month=TruncMonth('date_paiement')) \
        .values('month').annotate(total=Sum('montant')).order_by('month')

    labels_revenus = [p['month'].strftime("%b %Y") for p in paiements_mensuels]
    data_revenus = [float(p['total']) for p in paiements_mensuels]

    context = {
        'total_eleves': total_eleves,
        'total_paiements': total_paiements,
        'total_classes': total_classes,
        'moyenne_generale': round(moyenne_generale, 2),
        'labels_classes': json.dumps(labels_classes),
        'data_classes': json.dumps(data_classes),
        'labels_revenus': json.dumps(labels_revenus),
        'data_revenus': json.dumps(data_revenus),
    }
    return render(request, 'admin_dashboard.html', context)

@login_required
def responsable_dashboard(request):
    if request.user.role != "responsable":
        return redirect('admin_dashboard')
    total_eleves = Eleve.objects.count()
    return render(request, 'responsable_dashboard.html', {'total_eleves': total_eleves})

def deconnexion_view(request):
    logout(request)
    return redirect('login')

@login_required
def profile_view(request):
    return render(request, 'profile.html')

# --- GESTION DES CLASSES ---
@login_required
@user_passes_test(is_admin)
def liste_classes(request):
    classes = Classe.objects.all()
    return render(request, 'liste_classes.html', {'classes': classes})

@login_required
@user_passes_test(is_admin)
def ajouter_classe(request):
    if request.method == "POST":
        form = ClasseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('liste_classes')
    else:
        form = ClasseForm()
    return render(request, 'ajouter_classe.html', {'form': form})

# --- GESTION DES ÉLÈVES ---
@login_required
def liste_eleves(request):
    # Server-side filtering and pagination
    classe_id = request.GET.get('classe')
    q = request.GET.get('q', '').strip()

    eleves_qs = Eleve.objects.all().select_related('classe')
    if classe_id:
        eleves_qs = eleves_qs.filter(classe__id=classe_id)
    if q:
        eleves_qs = eleves_qs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(classe__nom__icontains=q)
        )

    paginator = Paginator(eleves_qs.order_by('nom', 'prenom'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    classes = Classe.objects.all()

    context = {
        'eleves': page_obj.object_list,
        'page_obj': page_obj,
        'classes': classes,
        'current_classe': classe_id,
        'q': q,
    }
    return render(request, 'liste_eleves.html', context)


@login_required
def export_eleves_csv(request):
    """Export the students list as CSV."""
    # Respect optional filtering query params
    classe_id = request.GET.get('classe')
    q = request.GET.get('q', '').strip()

    eleves = Eleve.objects.all().select_related('classe')
    if classe_id:
        eleves = eleves.filter(classe__id=classe_id)
    if q:
        eleves = eleves.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(classe__nom__icontains=q)
        )
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="eleves.csv"'

    writer = csv.writer(response)
    writer.writerow(['Nom', 'Prenom', 'Date de naissance', 'Classe', 'Contact'])
    for e in eleves:
        writer.writerow([e.nom, e.prenom, e.date_naissance, e.classe.nom if e.classe else '', getattr(e, 'contact', '')])

    return response


@login_required
def export_eleves_xlsx(request):
    classe_id = request.GET.get('classe')
    q = request.GET.get('q', '').strip()

    eleves = Eleve.objects.all().select_related('classe')
    if classe_id:
        eleves = eleves.filter(classe__id=classe_id)
    if q:
        eleves = eleves.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(classe__nom__icontains=q)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Élèves'
    ws.append(['Nom', 'Prenom', 'Date de naissance', 'Classe', 'Contact'])
    for e in eleves.order_by('nom', 'prenom'):
        ws.append([e.nom, e.prenom, str(e.date_naissance), e.classe.nom if e.classe else '', getattr(e, 'contact', '')])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="eleves.xlsx"'
    return response


@login_required
@user_passes_test(is_admin)
def export_responsables_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="responsables.csv"'
    writer = csv.writer(response)
    writer.writerow(['Username', 'Full name', 'Role', 'Sexe', 'Contact', 'Email', 'Is active'])
    responsables = Utilisateur.objects.filter(role='responsable')
    for r in responsables:
        writer.writerow([r.username, f"{r.first_name} {r.last_name}".strip(), r.role, r.sexe, r.contact, r.email, r.is_active])
    return response


@login_required
def export_responsables_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="responsables.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Responsables'
    ws.append(['Username', 'Full name', 'Role', 'Sexe', 'Contact', 'Email', 'Is active'])
    responsables = Utilisateur.objects.filter(role='responsable')
    for r in responsables:
        ws.append([r.username, f"{r.first_name} {r.last_name}".strip(), r.role, r.sexe, r.contact, r.email, r.is_active])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': 'attachment; filename="responsables.xlsx"'
    })


@login_required
def export_paiements_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="paiements.csv"'
    writer = csv.writer(response)
    writer.writerow(['Eleve', 'Classe', 'Montant', 'Date', 'Type', 'Commentaire'])

    paiements = Paiement.objects.select_related('eleve', 'eleve__classe')
    ptype = request.GET.get('type')
    start = request.GET.get('start')
    end = request.GET.get('end')
    if ptype:
        paiements = paiements.filter(type_paiement=ptype)
    if start:
        paiements = paiements.filter(date_paiement__gte=start)
    if end:
        paiements = paiements.filter(date_paiement__lte=end)

    for p in paiements.order_by('-date_paiement'):
        writer.writerow([f"{p.eleve.nom} {p.eleve.prenom}", p.eleve.classe.nom if p.eleve.classe else '', p.montant, p.date_paiement, p.get_type_paiement_display(), p.commentaire or ''])
    return response


@login_required
def export_paiements_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="paiements.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Paiements'
    ws.append(['Élève', 'Classe', 'Montant', 'Date', 'Type', 'Commentaire'])

    paiements = Paiement.objects.select_related('eleve', 'eleve__classe')
    ptype = request.GET.get('type')
    start = request.GET.get('start')
    end = request.GET.get('end')
    if ptype:
        paiements = paiements.filter(type_paiement=ptype)
    if start:
        paiements = paiements.filter(date_paiement__gte=start)
    if end:
        paiements = paiements.filter(date_paiement__lte=end)

    for p in paiements.order_by('-date_paiement'):
        ws.append([f"{p.eleve.nom} {p.eleve.prenom}", p.eleve.classe.nom if p.eleve.classe else '', p.montant, str(p.date_paiement), p.get_type_paiement_display(), p.commentaire or ''])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': 'attachment; filename="paiements.xlsx"'
    })


@login_required
def export_notes_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="notes.csv"'
    writer = csv.writer(response)
    writer.writerow(['Eleve', 'Classe', 'Matiere', 'Valeur', 'Date'])

    notes = Note.objects.select_related('eleve', 'eleve__classe')
    classe_id = request.GET.get('classe')
    start = request.GET.get('start')
    end = request.GET.get('end')
    if classe_id:
        notes = notes.filter(eleve__classe__id=classe_id)
    if start:
        notes = notes.filter(date_evaluation__gte=start)
    if end:
        notes = notes.filter(date_evaluation__lte=end)

    for n in notes.order_by('-date_evaluation'):
        writer.writerow([f"{n.eleve.nom} {n.eleve.prenom}", n.eleve.classe.nom if n.eleve.classe else '', n.matiere, n.valeur, n.date_evaluation])
    return response


@login_required
def export_notes_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="notes.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Notes'
    ws.append(['Élève', 'Classe', 'Matière', 'Valeur', 'Date'])

    notes = Note.objects.select_related('eleve', 'eleve__classe')
    classe_id = request.GET.get('classe')
    start = request.GET.get('start')
    end = request.GET.get('end')
    if classe_id:
        notes = notes.filter(eleve__classe__id=classe_id)
    if start:
        notes = notes.filter(date_evaluation__gte=start)
    if end:
        notes = notes.filter(date_evaluation__lte=end)

    for n in notes.order_by('-date_evaluation'):
        ws.append([f"{n.eleve.nom} {n.eleve.prenom}", n.eleve.classe.nom if n.eleve.classe else '', n.matiere, n.valeur, str(n.date_evaluation)])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': 'attachment; filename="notes.xlsx"'
    })


@login_required
def export_eleve_detail_csv(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    response = HttpResponse(content_type='text/csv')
    filename = f"eleve_{eleve.nom}_{eleve.prenom}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    # Header with personal info
    writer.writerow(['Eleve', eleve.nom, eleve.prenom])
    writer.writerow(['Date de naissance', eleve.date_naissance])
    writer.writerow(['Classe', eleve.classe.nom if eleve.classe else ''])
    writer.writerow([])

    # Payments
    writer.writerow(['Paiements'])
    writer.writerow(['Montant', 'Date', 'Type', 'Commentaire'])
    for p in eleve.paiements.all().order_by('-date_paiement'):
        writer.writerow([p.montant, p.date_paiement, p.get_type_paiement_display(), p.commentaire or ''])
    writer.writerow([])

    # Notes
    writer.writerow(['Notes'])
    writer.writerow(['Matiere', 'Valeur', 'Date'])
    for n in eleve.notes.all().order_by('-date_evaluation'):
        writer.writerow([n.matiere, n.valeur, n.date_evaluation])

    return response


@login_required
def export_eleve_detail_xlsx(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)
    wb = Workbook()
    info_ws = wb.active
    info_ws.title = 'Infos élève'
    info_ws.append(['Champ', 'Valeur'])
    info_ws.append(['Nom', eleve.nom])
    info_ws.append(['Prénom', eleve.prenom])
    info_ws.append(['Date de naissance', str(eleve.date_naissance)])
    info_ws.append(['Classe', eleve.classe.nom if eleve.classe else ''])

    paiements_ws = wb.create_sheet('Paiements')
    paiements_ws.append(['Montant', 'Date', 'Type', 'Commentaire'])
    for p in eleve.paiements.all().order_by('-date_paiement'):
        paiements_ws.append([p.montant, str(p.date_paiement), p.get_type_paiement_display(), p.commentaire or ''])

    notes_ws = wb.create_sheet('Notes')
    notes_ws.append(['Matière', 'Valeur', 'Date'])
    for n in eleve.notes.all().order_by('-date_evaluation'):
        notes_ws.append([n.matiere, n.valeur, str(n.date_evaluation)])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': f'attachment; filename="eleve_{eleve.nom}_{eleve.prenom}.xlsx"'
    })


@login_required
@user_passes_test(is_admin)
def export_responsable_detail_csv(request, user_id):
    user = get_object_or_404(Utilisateur, id=user_id, role='responsable')
    response = HttpResponse(content_type='text/csv')
    filename = f"responsable_{user.username}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(['Username', user.username])
    writer.writerow(['Nom', f"{user.first_name} {user.last_name}".strip()])
    writer.writerow(['Email', user.email])
    writer.writerow(['Contact', user.contact])
    writer.writerow(['Sexe', user.sexe])
    writer.writerow(['Role', user.role])
    return response


@login_required
@user_passes_test(is_admin)
def export_responsable_detail_xlsx(request, user_id):
    user = get_object_or_404(Utilisateur, id=user_id, role='responsable')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Responsable'
    ws.append(['Champ', 'Valeur'])
    ws.append(['Username', user.username])
    ws.append(['Nom', f"{user.first_name} {user.last_name}".strip()])
    ws.append(['Email', user.email])
    ws.append(['Contact', user.contact])
    ws.append(['Sexe', user.sexe])
    ws.append(['Role', user.role])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': f'attachment; filename="responsable_{user.username}.xlsx"'
    })


@login_required
def ajouter_eleve(request):
    if request.method == "POST":
        form = EleveForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('liste_eleves')
    else:
        form = EleveForm()
    return render(request, 'ajouter_eleve.html', {'form': form})

# --- GESTION DES PAIEMENTS ---
@login_required
def liste_paiements(request):
    type_selected = request.GET.get('type')
    start = request.GET.get('start')
    end = request.GET.get('end')

    paiements = Paiement.objects.select_related('eleve', 'eleve__classe').order_by('-date_paiement')
    if type_selected:
        paiements = paiements.filter(type_paiement=type_selected)
    if start:
        paiements = paiements.filter(date_paiement__gte=start)
    if end:
        paiements = paiements.filter(date_paiement__lte=end)

    return render(request, 'liste_paiements.html', {
        'paiements': paiements,
        'paiement_types': Paiement.TYPES,
        'type_selected': type_selected,
        'start': start,
        'end': end,
    })

@login_required
def ajouter_paiement(request):
    if request.method == "POST":
        form = PaiementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('liste_paiements')
    else:
        form = PaiementForm()
    return render(request, 'ajouter_paiement.html', {'form': form})

# --- GESTION DES NOTES ---
@login_required
def liste_notes(request):
    classe_id = request.GET.get('classe')
    start = request.GET.get('start')
    end = request.GET.get('end')

    notes = Note.objects.select_related('eleve', 'eleve__classe').order_by('-date_evaluation')
    if classe_id:
        notes = notes.filter(eleve__classe__id=classe_id)
    if start:
        notes = notes.filter(date_evaluation__gte=start)
    if end:
        notes = notes.filter(date_evaluation__lte=end)

    classes = Classe.objects.all()
    return render(request, 'liste_notes.html', {
        'notes': notes,
        'classes': classes,
        'classe_id': classe_id,
        'start': start,
        'end': end,
    })

@login_required
def ajouter_note(request):
    if request.method == "POST":
        form = NoteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('liste_notes')
    else:
        form = NoteForm()
    return render(request, 'ajouter_note.html', {'form': form})

# --- RAPPORTS ET STATISTIQUES ---
@login_required
@user_passes_test(is_admin)
def rapports_view(request):
    total_eleves = Eleve.objects.count()
    total_paiements = Paiement.objects.aggregate(Sum('montant'))['montant__sum'] or 0
    moyenne_generale = Note.objects.aggregate(Avg('valeur'))['valeur__avg'] or 0

    repartition_classe = Classe.objects.annotate(nb_eleves=Count('eleves'))
    derniers_paiements = Paiement.objects.all().order_by('-date_paiement')[:5]

    context = {
        'total_eleves': total_eleves,
        'total_paiements': total_paiements,
        'moyenne_generale': round(moyenne_generale, 2),
        'repartition_classe': repartition_classe,
        'derniers_paiements': derniers_paiements,
    }
    return render(request, 'rapports.html', context)


@login_required
@user_passes_test(is_admin)
def export_rapports_xlsx(request):
    total_eleves = Eleve.objects.count()
    total_paiements = Paiement.objects.aggregate(Sum('montant'))['montant__sum'] or 0
    moyenne_generale = Note.objects.aggregate(Avg('valeur'))['valeur__avg'] or 0
    repartition_classe = Classe.objects.annotate(nb_eleves=Count('eleves'))
    derniers_paiements = Paiement.objects.all().order_by('-date_paiement')[:20]

    wb = Workbook()
    summary = wb.active
    summary.title = 'Résumé'
    summary.append(['Clé', 'Valeur'])
    summary.append(['Total élèves', total_eleves])
    summary.append(['Recettes totales', total_paiements])
    summary.append(['Moyenne générale', round(moyenne_generale, 2)])

    classe_ws = wb.create_sheet('Répartition classes')
    classe_ws.append(['Classe', 'Niveau', 'Nombre d’élèves'])
    for c in repartition_classe:
        classe_ws.append([c.nom, c.niveau, c.nb_eleves])

    paiement_ws = wb.create_sheet('Derniers paiements')
    paiement_ws.append(['Élève', 'Classe', 'Montant', 'Date', 'Type', 'Commentaire'])
    for p in derniers_paiements:
        paiement_ws.append([
            f"{p.eleve.nom} {p.eleve.prenom}",
            p.eleve.classe.nom if p.eleve.classe else '',
            p.montant,
            str(p.date_paiement),
            p.get_type_paiement_display(),
            p.commentaire or ''
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': 'attachment; filename="rapports_campus.xlsx"'
    })
